import os
import sys
import openpyxl

# Add project root to path
sys.path.append(r'C:\Users\njuku\Documents\AI COMPANY\SephlightyAI\backend')

try:
    from laravel_modules.ai_brain.omnibrain_saas_engine import OmnibrainSaaSEngine
    brain = OmnibrainSaaSEngine()
    
    # 1. Trigger Export
    res = brain.process_query('export sales to excel', 'test_conn')
    response_str = res.get('response', '')
    
    if '[DOWNLOAD_ACTION]' in response_str:
        url = response_str.split('[DOWNLOAD_ACTION]: ')[1].strip()
        filename = url.split('/')[-1]
        
        path = os.path.join('exports', filename)
        if os.path.exists(path):
            # Verify sheets
            wb = openpyxl.load_workbook(path)
            
            if 'Data_Report' in wb.sheetnames and 'AI_STRATEGY' in wb.sheetnames:
                # Check styling on first sheet
                ws = wb['Data_Report']
                if ws.freeze_panes == 'A2':
                    sys.exit(0) # SUCCESS
                else:
                    sys.exit(2) # STYLING ERROR
            else:
                sys.exit(3) # SHEET ERROR
        else:
            sys.exit(4) # FILE NOT FOUND
    else:
        sys.exit(5) # TAG ERROR

except Exception as e:
    sys.exit(6) # GENERAL ERROR
