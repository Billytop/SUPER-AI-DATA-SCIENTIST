import os
import sys
import openpyxl

# Add project root to path
sys.path.append(r'C:\Users\njuku\Documents\AI COMPANY\SephlightyAI\backend')

try:
    from laravel_modules.ai_brain.omnibrain_saas_engine import OmnibrainSaaSEngine
    brain = OmnibrainSaaSEngine()
    
    # 1. Trigger Export with TYPO
    query = 'export mauzo ya mwenzi huu to excel'
    print(f"Executing query: {query}")
    res = brain.process_query(query, 'test_conn')
    response_str = res.get('response', '')
    
    # Avoid printing emojis
    clean_response = response_str.encode('ascii', 'ignore').decode('ascii')
    print(f"Response: {clean_response[:100]}...")
    
    if '[DOWNLOAD_ACTION]' in response_str:
        url = response_str.split('[DOWNLOAD_ACTION]: ')[1].strip()
        filename = url.split('/')[-1]
        
        # Search for the file
        possible_paths = [
            os.path.abspath(os.path.join('exports', filename)),
            os.path.abspath(os.path.join('laravel_modules', 'ai_brain', 'exports', filename)),
            os.path.abspath(os.path.join('backend', 'exports', filename))
        ]
        
        found_path = None
        for p in possible_paths:
            if os.path.exists(p):
                found_path = p
                break
        
        if found_path:
            print(f"SUCCESS: File generated at {found_path}")
            
            # Verify sheets
            wb = openpyxl.load_workbook(found_path)
            print(f"Sheet names: {wb.sheetnames}")
            
            if 'Data_Report' in wb.sheetnames:
                ws = wb['Data_Report']
                print(f"SUCCESS: Data_Report sheet found with {ws.max_column} columns.")
                sys.exit(0)
            else:
                print("ERROR: Data_Report sheet missing.")
                sys.exit(2)
        else:
            print(f"ERROR: File not found in {possible_paths}")
            sys.exit(3)
    else:
        print(f"ERROR: [DOWNLOAD_ACTION] tag missing. Response was: {clean_response}")
        sys.exit(4)

except Exception as e:
    print(f"Error: {e}")
    sys.exit(5)
