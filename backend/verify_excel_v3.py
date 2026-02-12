import os
import sys
import openpyxl

# Add project root to path
sys.path.append(r'C:\Users\njuku\Documents\AI COMPANY\SephlightyAI\backend')

try:
    from laravel_modules.ai_brain.omnibrain_saas_engine import OmnibrainSaaSEngine
    brain = OmnibrainSaaSEngine()
    
    # 1. Trigger Export
    res = brain.process_query('export sales to excel', 'test_conn')
    response_str = res.get('response', '')
    
    # Avoid printing emojis to terminal
    clean_response = response_str.encode('ascii', 'ignore').decode('ascii')
    
    if '[DOWNLOAD_ACTION]' in response_str:
        url = response_str.split('[DOWNLOAD_ACTION]: ')[1].strip()
        filename = url.split('/')[-1]
        
        # Search for the file in possible locations
        possible_paths = [
            os.path.abspath(os.path.join('exports', filename)),
            os.path.abspath(os.path.join('laravel_modules', 'ai_brain', 'exports', filename)),
            os.path.abspath(os.path.join('backend', 'exports', filename))
        ]
        
        found_path = None
        for p in possible_paths:
            if os.path.exists(p):
                found_path = p
                break
        
        if found_path:
            print(f"SUCCESS: File generated at {found_path}")
            
            # Verify sheets
            wb = openpyxl.load_workbook(found_path)
            print(f"Sheet names: {wb.sheetnames}")
            
            if 'Data_Report' in wb.sheetnames and 'AI_STRATEGY' in wb.sheetnames:
                print("SUCCESS: Both sheets found.")
                
                # Check styling on first sheet
                ws = wb['Data_Report']
                if ws.freeze_panes == 'A2':
                    print("SUCCESS: Header frozen (A2).")
                    sys.exit(0)
                else:
                    print(f"ERROR: Styling mismatch (freeze_panes={ws.freeze_panes})")
                    sys.exit(2)
            else:
                print(f"ERROR: Missing sheets. Found: {wb.sheetnames}")
                sys.exit(3)
        else:
            print(f"ERROR: File not found in {possible_paths}")
            sys.exit(4)
    else:
        print("ERROR: [DOWNLOAD_ACTION] tag missing.")
        sys.exit(5)

except Exception as e:
    print(f"Error: {e}")
    sys.exit(6)
